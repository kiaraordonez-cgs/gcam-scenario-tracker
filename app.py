#!/usr/bin/env python3
"""
GCAM Scenario Tracker - Google Drive/Sheets Version
A Flask app that stores data in Google Sheets and files in Google Drive
"""

import os
import json
import difflib
import uuid
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response
from werkzeug.utils import secure_filename
from lxml import etree
import gspread
from google.oauth2.service_account import Credentials

# Load variables from a local .env file for development.
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# =============================================================================
# Configuration
# =============================================================================

app = Flask(__name__)
app.config['SECRET_KEY'] = 'change-this-to-a-random-secret-key'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max



def _parse_timestamp(value):
    """Parse the timestamp formats found in the sheet. None if unrecognised."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    iso = text.replace(' ', 'T', 1)
    try:
        return datetime.fromisoformat(iso)
    except ValueError:
        pass
    for fmt in ('%m/%d/%Y %H:%M:%S', '%m/%d/%Y %H:%M', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


@app.template_filter('humandatetime')
def humandatetime(value):
    """2026-07-14T13:23:10 -> Jul 14, 2026, 1:23 PM"""
    parsed = _parse_timestamp(value)
    if parsed is None:
        return '' if value is None else str(value)
    hour = parsed.hour % 12 or 12
    meridiem = 'AM' if parsed.hour < 12 else 'PM'
    return f"{parsed:%b} {parsed.day}, {parsed.year}, {hour}:{parsed.minute:02d} {meridiem}"


@app.template_filter('humandate')
def humandate(value):
    """2026-07-14T13:23:10 -> Jul 14, 2026"""
    parsed = _parse_timestamp(value)
    if parsed is None:
        return '' if value is None else str(value)
    return f"{parsed:%b} {parsed.day}, {parsed.year}"

# Google Configuration
SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]

GOOGLE_SHEET_ID = os.environ.get('GOOGLE_SHEET_ID', '1n1dwcHThv_I19lWkuNhAK5rlscnyMGlM2x-VoR4R8rs')

# =============================================================================
# Naming Convention Mappings (loaded from CSV files)
# =============================================================================
# Filename pattern: configuration_PROJECT_scenarioAbbrev_YYMMDD-FL_comment.xml

import csv

def load_csv_mapping(filepath, key_col, value_col):
    """Load a CSV file into a dict mapping key_col -> value_col"""
    mapping = {}
    try:
        with open(filepath, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                k = row.get(key_col, '').strip()
                v = row.get(value_col, '').strip()
                if k and v:
                    mapping[k] = v
    except FileNotFoundError:
        print(f"Warning: Mapping file not found: {filepath}")
    except Exception as e:
        print(f"Warning: Error loading {filepath}: {e}")
    return mapping

# Load mappings from CSV files (in mappings/ directory next to app.py)
MAPPINGS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mappings')
PROJECT_MAP = load_csv_mapping(os.path.join(MAPPINGS_DIR, 'projects.csv'), 'code', 'name')
PERSON_MAP = load_csv_mapping(os.path.join(MAPPINGS_DIR, 'people.csv'), 'initials', 'name')

USERNAME_MAP = load_csv_mapping(os.path.join(MAPPINGS_DIR, 'people.csv'), 'zaratan_username', 'name')

print(f"Loaded {len(PROJECT_MAP)} project mappings: {PROJECT_MAP}")
print(f"Loaded {len(PERSON_MAP)} person mappings: {PERSON_MAP}")
print(f"Loaded {len(USERNAME_MAP)} zaratan username mappings: {USERNAME_MAP}")

def parse_scenario_filename(filename):
    """Parse configuration filename to extract metadata.
    
    Pattern: configuration_PROJECT_abbrev_YYMMDD-FL_comment.xml
    Returns dict with extracted fields, or empty dict if pattern doesn't match.
    """
    result = {}
    
    name = filename
    if name.lower().startswith('configuration_'):
        name = name[14:]
    if name.lower().endswith('.xml'):
        name = name[:-4]
    if not name:
        return result
    
    parts = name.split('_')
    if len(parts) < 3:
        return result
    
    # First part: project code
    project_code = parts[0]
    if project_code in PROJECT_MAP:
        result['project_name'] = PROJECT_MAP[project_code]
        result['project_code'] = project_code
    
    # Find the part with YYMMDD-FL (contains a dash with 6-digit date)
    date_person_idx = None
    for i, part in enumerate(parts):
        if '-' in part:
            segments = part.split('-')
            if len(segments) >= 2 and segments[0].isdigit() and len(segments[0]) == 6:
                date_person_idx = i
                break
    
    if date_person_idx is not None:
        segments = parts[date_person_idx].split('-')
        
        # Parse date: YYMMDD
        date_str = segments[0]
        try:
            year = int('20' + date_str[:2])
            month = int(date_str[2:4])
            day = int(date_str[4:6])
            result['date_run'] = f'{year}-{month:02d}-{day:02d}'
        except (ValueError, IndexError):
            pass
        
        # Parse initials
        if len(segments) >= 2:
            initials = segments[1].upper()
            if initials in PERSON_MAP:
                result['person_name'] = PERSON_MAP[initials]
                result['person_initials'] = initials
        
        # Scenario abbreviation: everything between project code and date
        abbrev_parts = parts[1:date_person_idx]
        if abbrev_parts:
            result['scenario_abbrev'] = '_'.join(abbrev_parts)
        
        # Comment: everything after date-initials
        comment_parts = parts[date_person_idx + 1:]
        if comment_parts:
            result['comment'] = '_'.join(comment_parts)
    else:
        result['scenario_abbrev'] = '_'.join(parts[1:])
    
    return result
# Google Drive (Shared Drive "GCAM Scenario Tracker Files").
GOOGLE_DRIVE_FOLDER_ID = os.environ.get('GOOGLE_DRIVE_FOLDER_ID', '1MHMq9vN6QFktXvfo33tG2OpUlcbvblSg')
DRIVE_CONFIGS_FOLDER_ID = os.environ.get('DRIVE_CONFIGS_FOLDER_ID', '1oXhLFZfEkd-HFKae5Og9c-0ubrFiqAiW')
DRIVE_INPUTS_FOLDER_ID = os.environ.get('DRIVE_INPUTS_FOLDER_ID', '1Q85c9N6N6YvfeugKzr3MdqxpfK-jhAnw')

ALLOWED_EXTENSIONS = {'xml'}

# Zaratan log ingestion token for security (Optional)
INGEST_TOKEN = os.environ.get('INGEST_TOKEN', '')

# Column order for the ZaratanLogs worksheet (raw log storage).
ZARATAN_LOG_COLUMNS = [
    'job_id', 'restart_count', 'event', 'scenario_name', 'project_name', 'user',
    'hostname', 'started_at', 'config_file', 'gcam_dir', 'gcamreport_template_version',
    'gcam_finished_at', 'gcam_exit_code', 'script_finished_at', 'script_exit_code',
    'sacct_submit', 'sacct_start', 'sacct_end', 'sacct_elapsed', 'sacct_state',
    'sacct_exit_code', 'received_at', 'raw_json'
]

# =============================================================================
# Google API Setup
# =============================================================================

def get_google_sheets_client():
    """Initialize and return Google Sheets client"""
    import os
    import json
    import base64
    
    # Try base64 encoded env var first
    creds_base64 = os.environ.get('GOOGLE_CREDENTIALS_BASE64')
    if creds_base64:
        print("DEBUG: Using base64 encoded credentials")
        creds_json = base64.b64decode(creds_base64).decode('utf-8')
        creds_dict = json.loads(creds_json)
        creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    else:
        # Fallback to JSON env var
        creds_json = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS_JSON')
        if creds_json:
            creds_dict = json.loads(creds_json)
            creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
        else:
            creds = Credentials.from_service_account_file('service-account.json', scopes=SCOPES)
    
    return gspread.authorize(creds)


_drive_service = None

def get_drive_service():
    """Lazily build the Drive client, caching it for the process.

    Returns None on failure rather than raising, so a Drive problem degrades
    the affected feature instead of taking down the whole app.
    """
    global _drive_service
    if _drive_service is not None:
        return _drive_service
    try:
        import base64
        from googleapiclient.discovery import build

        creds_base64 = os.environ.get('GOOGLE_CREDENTIALS_BASE64')
        creds_json = os.environ.get('GOOGLE_APPLICATION_CREDENTIALS_JSON')
        if creds_base64:
            info = json.loads(base64.b64decode(creds_base64).decode('utf-8'))
            creds = Credentials.from_service_account_info(info, scopes=SCOPES)
        elif creds_json:
            creds = Credentials.from_service_account_info(json.loads(creds_json), scopes=SCOPES)
        else:
            creds = Credentials.from_service_account_file('service-account.json', scopes=SCOPES)

        _drive_service = build('drive', 'v3', credentials=creds, cache_discovery=False)
        return _drive_service
    except Exception as e:
        print(f"Error initializing Drive client: {e}")
        return None


def drive_available():
    return get_drive_service() is not None


def upload_file_to_drive(content, filename, folder_id=None, mime_type='application/xml'):
    """Store text content in Drive and return the new file's id, or None.

    A repeat upload of the same filename creates a second file rather than
    replacing the first. That is intentional: each scenario keeps the exact
    config it was uploaded with, even when two share a name.
    """
    service = get_drive_service()
    if service is None:
        return None
    try:
        from io import BytesIO
        from googleapiclient.http import MediaIoBaseUpload

        data = content.encode('utf-8') if isinstance(content, str) else content
        media = MediaIoBaseUpload(BytesIO(data), mimetype=mime_type, resumable=False)
        metadata = {'name': filename, 'parents': [folder_id or DRIVE_CONFIGS_FOLDER_ID]}
        created = service.files().create(
            body=metadata, media_body=media, fields='id', supportsAllDrives=True
        ).execute()
        return created.get('id')
    except Exception as e:
        print(f"Error uploading {filename} to Drive: {e}")
        return None


def download_file_from_drive(file_id):
    """Return a Drive file's contents as text, or None if unavailable."""
    service = get_drive_service()
    if service is None:
        return None
    try:
        data = service.files().get_media(fileId=file_id, supportsAllDrives=True).execute()
        return data.decode('utf-8', 'replace') if isinstance(data, bytes) else data
    except Exception as e:
        print(f"Error downloading {file_id} from Drive: {e}")
        return None


def is_legacy_file_id(file_id):
    """Files stored before the Drive migration live in the FileStorage sheet and
    are keyed by a small integer; Drive ids are long opaque strings."""
    return str(file_id).strip().isdigit()


def get_file_content(file_id):
    """Read a stored file from whichever backend holds it.

    Keeps the two storage generations working side by side so the migration
    does not have to be atomic.
    """
    if file_id is None or not str(file_id).strip():
        return None
    if is_legacy_file_id(file_id):
        return download_file_from_sheet(file_id)
    return download_file_from_drive(file_id)

# Initialize clients
try:
    gc = get_google_sheets_client()
    sheet = gc.open_by_key(GOOGLE_SHEET_ID)
    scenarios_sheet = sheet.worksheet('Scenarios')
    inputs_sheet = sheet.worksheet('InputFiles')
    junction_sheet = sheet.worksheet('ScenarioInputs')
    
    # FileStorage sheet for storing XML content
    try:
        file_storage_sheet = sheet.worksheet('FileStorage')
    except:
        # Create if doesn't exist
        file_storage_sheet = sheet.add_worksheet(title='FileStorage', rows=1000, cols=3)
        file_storage_sheet.append_row(['file_id', 'filename', 'content'])

    # ZaratanLogs sheet for storing raw run logs sent from the cluster
    try:
        zaratan_logs_sheet = sheet.worksheet('ZaratanLogs')
    except:
        zaratan_logs_sheet = sheet.add_worksheet(title='ZaratanLogs', rows=1000, cols=len(ZARATAN_LOG_COLUMNS))
        zaratan_logs_sheet.append_row(ZARATAN_LOG_COLUMNS)

    print("✓ Google Sheets connection successful")
except Exception as e:
    print(f"✗ Error initializing Google APIs: {e}")
    print("⚠ App will start but Google Sheets features will be unavailable")
    gc = None
    scenarios_sheet = None
    inputs_sheet = None
    junction_sheet = None
    file_storage_sheet = None
    zaratan_logs_sheet = None

# =============================================================================
# Helper Functions - Google Sheets
# =============================================================================

def sheets_available():
    """Check if Google Sheets connection is available"""
    return scenarios_sheet is not None


SCENARIO_TEXT_COLUMNS = [1, 11, 16]   # id, config_file_id, job_id
INPUT_TEXT_COLUMNS = [1, 12]          # id, file_id
JUNCTION_TEXT_COLUMNS = [1, 2]        # scenario_id, input_file_id

def read_scenario_records():
    return scenarios_sheet.get_all_records(numericise_ignore=SCENARIO_TEXT_COLUMNS)

def read_input_records():
    return inputs_sheet.get_all_records(numericise_ignore=INPUT_TEXT_COLUMNS)

def read_junction_records():
    return junction_sheet.get_all_records(numericise_ignore=JUNCTION_TEXT_COLUMNS)

# Simple in-memory cache to speed up page loads
_cache = {'data': None, 'timestamp': 0}
CACHE_TTL = 60  # seconds - longer cache = fewer API calls

def get_cached_data():
    """Get all sheet data with caching to reduce API calls"""
    import time
    now = time.time()
    if _cache['data'] and (now - _cache['timestamp']) < CACHE_TTL:
        return _cache['data']
    
    scenarios = []
    junctions = []
    inputs = []
    
    try:
        print("Loading scenarios...")
        scenarios = read_scenario_records()
        print(f"  Loaded {len(scenarios)} scenarios")
    except Exception as e:
        print(f"  ERROR loading scenarios: {e}")
    
    try:
        print("Loading junctions...")
        junctions = read_junction_records()
        print(f"  Loaded {len(junctions)} junctions")
    except Exception as e:
        print(f"  ERROR loading junctions: {e}")
    
    try:
        print("Loading input files...")
        inputs = read_input_records()
        print(f"  Loaded {len(inputs)} input files")
    except Exception as e:
        print(f"  ERROR loading inputs: {e}")
    
    _cache['data'] = {'scenarios': scenarios, 'junctions': junctions, 'inputs': inputs}
    _cache['timestamp'] = now
    return _cache['data']

def invalidate_cache():
    """Clear cache after writes"""
    _cache['data'] = None
    _cache['timestamp'] = 0

def patch_cached_record(collection, record_id, updates):
    """Apply field updates to a record already held in the cache.

    Invalidating after a single-field edit is expensive: the next /api/data
    then cold-reads all three sheets. Since we know exactly what changed, we
    patch the cached record in place instead and leave the TTL as a backstop.

    'updates' keys are sheet header names, which are also the keys produced by
    get_all_records() and the field names posted by the dashboard.

    Returns the patched record, or None when the cache is cold or the record
    is absent — callers must then fall back to invalidate_cache(), since a
    partial cache is worse than no cache.

    Safe without locking under the deployed config (gunicorn, one sync worker;
    see Procfile). Revisit if the app ever runs multi-threaded.
    """
    data = _cache.get('data')
    if not data:
        return None
    for record in data.get(collection, []):
        if str(record.get('id')) == str(record_id):
            record.update(updates)
            return record
    return None

def compute_duration(submitted_str, finished_str):
    """Elapsed time between two stored timestamps, as e.g. '3h 12m'.

    The two ends of a run are not stored in the same format - 'submitted' is
    typically naive ('2026-07-14 13:23:10') while 'finished' carries a UTC
    offset ('2026-07-14T13:43:24-04:00'). Each side is therefore parsed
    independently; requiring a single format to fit both is what made this
    return '' for every record in the sheet.

    Returns '' when either side is unparseable or the range is negative.
    """
    start = _parse_timestamp(submitted_str)
    end = _parse_timestamp(finished_str)
    if start is None or end is None:
        return ''
    if (start.tzinfo is None) != (end.tzinfo is None):
        if start.tzinfo is None:
            start = start.replace(tzinfo=end.tzinfo)
        else:
            end = end.replace(tzinfo=start.tzinfo)

    total_minutes = int((end - start).total_seconds() // 60)
    if total_minutes < 0:
        return ''
    hours, minutes = divmod(total_minutes, 60)
    return f"{hours}h {minutes}m" if hours else f"{minutes}m"

def get_next_id(worksheet):
    """Get next available ID for a sheet"""
    values = worksheet.col_values(1)[1:]  # Skip header
    if not values:
        return 1
    return max([int(v) for v in values if v.isdigit()]) + 1

def find_row_by_id(worksheet, row_id):
    """Find row number by ID (returns row number, 1-indexed)"""
    try:
        cell = worksheet.find(str(row_id), in_column=1)
        return cell.row if cell else None
    except:
        return None

def find_row_by_value(worksheet, column, value):
    """Find first row where column matches value"""
    try:
        col_idx = worksheet.find(column).col
        col_values = worksheet.col_values(col_idx)
        for idx, val in enumerate(col_values[1:], start=2):  # Skip header
            if val == value:
                return idx
        return None
    except:
        return None

def get_all_scenarios():
    """Get all scenarios from Google Sheet"""
    try:
        records = read_scenario_records()
        # Count input files for each scenario
        junction_records = read_junction_records()
        
        for record in records:
            record['input_count'] = sum(1 for j in junction_records if str(j.get('scenario_id')) == str(record.get('id')))
        
        return records
    except Exception as e:
        print(f"Error getting scenarios: {e}")
        return []

def get_all_input_files():
    """Get all input files from Google Sheet"""
    try:
        records = read_input_records()
        # Count scenarios for each input file
        junction_records = read_junction_records()
        
        for record in records:
            record['scenario_count'] = sum(1 for j in junction_records if str(j.get('input_file_id')) == str(record.get('id')))
        
        return records
    except Exception as e:
        print(f"Error getting input files: {e}")
        return []

def get_scenario_by_id(scenario_id):
    """Get scenario by ID"""
    try:
        records = read_scenario_records()
        for record in records:
            if str(record.get('id')) == str(scenario_id):
                return record
        return None
    except Exception as e:
        print(f"Error getting scenario: {e}")
        return None

def get_input_by_id(input_id):
    """Get input file by ID"""
    try:
        records = read_input_records()
        for record in records:
            if str(record.get('id')) == str(input_id):
                return record
        return None
    except Exception as e:
        print(f"Error getting input file: {e}")
        return None

def get_input_files_for_scenario(scenario_id):
    """Get all input files linked to a scenario"""
    try:
        junction_records = read_junction_records()
        input_records = read_input_records()
        
        linked_input_ids = [j['input_file_id'] for j in junction_records if str(j['scenario_id']) == str(scenario_id)]
        
        result = []
        for input_rec in input_records:
            if str(input_rec['id']) in [str(i) for i in linked_input_ids]:
                # Find component key
                for j in junction_records:
                    if str(j['scenario_id']) == str(scenario_id) and str(j['input_file_id']) == str(input_rec['id']):
                        input_rec['component_key'] = j.get('component_key', '')
                        break
                result.append(input_rec)
        
        return result
    except Exception as e:
        print(f"Error getting input files for scenario: {e}")
        return []

def get_scenarios_for_input(input_id):
    """Get all scenarios using an input file"""
    try:
        junction_records = read_junction_records()
        scenario_records = read_scenario_records()
        
        linked_scenario_ids = [j['scenario_id'] for j in junction_records if str(j['input_file_id']) == str(input_id)]
        
        result = []
        for scenario_rec in scenario_records:
            if str(scenario_rec['id']) in [str(s) for s in linked_scenario_ids]:
                # Find component key
                for j in junction_records:
                    if str(j['input_file_id']) == str(input_id) and str(j['scenario_id']) == str(scenario_rec['id']):
                        scenario_rec['component_key'] = j.get('component_key', '')
                        break
                result.append(scenario_rec)
        
        return result
    except Exception as e:
        print(f"Error getting scenarios for input: {e}")
        return []

# =============================================================================
# Helper Functions - File Storage (in Sheets)
# =============================================================================

def upload_file_to_sheet(file_content, filename):
    """Store file content in Google Sheets and return file ID"""
    try:
        file_id = get_next_id(file_storage_sheet)
        file_storage_sheet.append_row([
            file_id,
            filename,
            file_content
        ])
        print(f"DEBUG: Stored file {filename} in Sheets, ID: {file_id}")
        return str(file_id)
    except Exception as e:
        print(f"Error storing file in Sheets: {e}")
        return None

def download_file_from_sheet(file_id):
    """Download file content from Google Sheets"""
    try:
        # Find row with matching file_id
        cell = file_storage_sheet.find(str(file_id), in_column=1)
        if cell:
            row_data = file_storage_sheet.row_values(cell.row)
            return row_data[2] if len(row_data) > 2 else None
        return None
    except Exception as e:
        print(f"Error downloading file from Sheets: {e}")
        return None

# =============================================================================
# XML Parsing Functions
# =============================================================================

def parse_configuration_xml(xml_content):
    """Parse GCAM configuration XML from string content"""
    try:
        root = etree.fromstring(xml_content.encode())
    except:
        root = etree.fromstring(xml_content)
    
    result = {
        'scenario_name': None,
        'input_files': []
    }
    
    # Extract scenario name
    for val in root.findall('.//Strings/Value[@name="scenarioName"]'):
        if val.text:
            result['scenario_name'] = val.text.strip()
            break
    
    if not result['scenario_name']:
        result['scenario_name'] = 'Unnamed Scenario'
    
    # Extract input files
    for comp in root.findall('.//ScenarioComponents/Value'):
        file_path = comp.text.strip() if comp.text else None
        component_key = comp.get('name', '')
        
        if file_path:
            file_name = Path(file_path).name
            
            # Extract folder location - everything between "input/" and the filename
            # Examples: 
            #   ../input/gcamdata/xml/file.xml → gcamdata/xml
            #   ../input/policyAI/file.xml → policyAI
            #   input/policyAI/file.xml → policyAI
            #   input/magicc/inputs/file.emk → magicc/inputs
            folder_location = ''
            if '/input/' in file_path:
                # Handle paths like ../input/xxx or /input/xxx
                parts = file_path.split('/input/', 1)
                if len(parts) > 1:
                    after_input = parts[1]
                    folder_parts = after_input.split('/')[:-1]  # Remove filename
                    folder_location = '/'.join(folder_parts)
            elif file_path.startswith('input/'):
                # Handle paths starting with input/
                after_input = file_path[6:]  # Remove 'input/'
                folder_parts = after_input.split('/')[:-1]  # Remove filename
                folder_location = '/'.join(folder_parts)
            
            result['input_files'].append({
                'file_name': file_name,
                'file_path': file_path,
                'folder_location': folder_location,
                'component_key': component_key
            })
    
    return result

def parse_input_file_xml(xml_content):
    """Parse GCAM input file XML from string content"""
    try:
        try:
            root = etree.fromstring(xml_content.encode())
        except:
            root = etree.fromstring(xml_content)
        
        result = {
            'regions': set(),
            'years': set(),
            'sectors': set()
        }
        
        # Extract regions
        for region in root.findall('.//region'):
            region_name = region.get('name')
            if region_name:
                result['regions'].add(region_name)
        
        # Extract years
        for constraint in root.findall('.//constraint'):
            year = constraint.get('year')
            if year:
                result['years'].add(year)
        
        # Extract sectors/policies
        for policy in root.findall('.//policy-portfolio-standard'):
            policy_name = policy.get('name', '')
            if policy_name:
                result['sectors'].add(policy_name)
        
        return {
            'regions': ', '.join(sorted(result['regions'])) if result['regions'] else 'All',
            'years': ', '.join(sorted(result['years'])) if result['years'] else 'N/A',
            'sectors': ', '.join(sorted(result['sectors'])) if result['sectors'] else 'N/A'
        }
    except Exception as e:
        print(f"Error parsing input XML: {e}")
        return {
            'regions': 'Parse Error',
            'years': 'Parse Error',
            'sectors': 'Parse Error'
        }

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# =============================================================================
# Routes
# =============================================================================

@app.route('/health')
def health():
    """Health check endpoint for Render"""
    status = {
        'status': 'ok',
        'sheets_connected': sheets_available()
    }
    return jsonify(status)

@app.route('/')
def index():
    """Main dashboard - loads instantly, data fetched via AJAX"""
    return render_template('index.html',
                         scenarios=[],
                         input_files=[],
                         projects=[])

@app.route('/api/mappings')
def api_mappings():
    """View current project and person mappings"""
    return jsonify({
        'projects': PROJECT_MAP,
        'people': PERSON_MAP,
        'filename_pattern': 'configuration_PROJECT_scenarioAbbrev_YYMMDD-FL_comment.xml',
        'example': 'configuration_USAI_HA1_250206-KO_GDPPOPhigh.xml'
    })

@app.route('/api/health')
def api_health():
    """Diagnostic endpoint - check what's working and what's not"""
    import time
    results = {'sheets_available': sheets_available()}
    
    if not sheets_available():
        return jsonify(results)
    
    for name, sheet in [('Scenarios', scenarios_sheet), ('InputFiles', inputs_sheet), ('ScenarioInputs', junction_sheet)]:
        start = time.time()
        try:
            rows = sheet.get_all_records()
            elapsed = round(time.time() - start, 2)
            results[name] = {'status': 'ok', 'rows': len(rows), 'seconds': elapsed}
        except Exception as e:
            elapsed = round(time.time() - start, 2)
            results[name] = {'status': 'error', 'error': str(e), 'seconds': elapsed}
    
    return jsonify(results)

@app.route('/api/data')
def api_data():
    """JSON endpoint for dashboard data - called via AJAX after page loads"""
    if not sheets_available():
        return jsonify({'error': 'Sheets unavailable'}), 503
    
    cached = get_cached_data()
    if not cached:
        return jsonify({'error': 'Could not load data'}), 500
    
    scenarios = cached['scenarios']
    junction_records = cached['junctions']
    input_files = cached['inputs']
    
    # Build set of valid scenario IDs to filter out orphaned junctions
    valid_scenario_ids = set(str(s.get('id')) for s in scenarios)
    
    # Pre-build count dicts for O(1) lookups instead of O(n) scans per record
    scenario_input_counts = {}
    input_scenario_counts = {}
    for j in junction_records:
        sid = str(j.get('scenario_id', ''))
        iid = str(j.get('input_file_id', ''))
        if sid in valid_scenario_ids:
            scenario_input_counts[sid] = scenario_input_counts.get(sid, 0) + 1
            input_scenario_counts[iid] = input_scenario_counts.get(iid, 0) + 1
    
    for record in scenarios:
        record['input_count'] = scenario_input_counts.get(str(record.get('id')), 0)
        submitted = str(record.get('submitted', ''))
        finished = str(record.get('finished', ''))
        if submitted and finished:
            record['duration'] = compute_duration(submitted, finished)
        else:
            record['duration'] = ''
    
    for record in input_files:
        record['scenario_count'] = input_scenario_counts.get(str(record.get('id')), 0)
    
    projects = sorted(set([s.get('project_name', '') for s in scenarios if s.get('project_name')]))
    
    return jsonify({
        'scenarios': scenarios,
        'input_files': input_files,
        'projects': projects
    })

def link_config_inputs(scenario_id, input_files, upload_date, uploaded_by='Auto-detected'):
    """Create InputFiles rows for a config's components and link them to a
    scenario. Shared by the website upload path and Zaratan log ingestion so
    both produce identical data.

    Batched into at most two append calls regardless of how many components a
    config declares - one config can reference ~180 files, and a per-row write
    would exhaust the Sheets quota immediately.

    Returns {'linked': N, 'created': M}.
    """
    if not input_files:
        return {'linked': 0, 'created': 0}

    existing_files = {}
    try:
        for inp in read_input_records():
            existing_files[inp['file_name']] = inp['id']
    except Exception as e:
        print(f"WARN: could not read existing input files: {e}")

    new_files_to_add = []
    junctions_to_add = []
    next_input_id = get_next_id(inputs_sheet)

    for input_file in input_files:
        name = input_file['file_name']
        if name in existing_files:
            input_id = existing_files[name]
        else:
            input_id = next_input_id
            new_files_to_add.append([
                input_id,
                name,
                'Not analyzed',
                'Not analyzed',
                'Not analyzed',
                '',  # policy_name
                input_file.get('folder_location', ''),
                '',  # description
                '',  # additional_notes
                uploaded_by,
                upload_date,
                ''   # file_id - set once the file itself is archived
            ])
            existing_files[name] = input_id
            next_input_id += 1

        junctions_to_add.append([scenario_id, input_id, input_file['component_key']])

    if new_files_to_add:
        inputs_sheet.append_rows(new_files_to_add)
    if junctions_to_add:
        junction_sheet.append_rows(junctions_to_add)

    return {'linked': len(junctions_to_add), 'created': len(new_files_to_add)}


@app.route('/upload_config', methods=['POST'])
def upload_config():
    """Upload and parse configuration XML"""
    if 'config_file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('index'))
    
    file = request.files['config_file']
    uploaded_by = 'Website'  # Fixed - website uploads are always "Website"
    
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    if not allowed_file(file.filename):
        flash('Only XML files allowed', 'error')
        return redirect(url_for('index'))
    
    try:
        # Read file content
        filename = secure_filename(file.filename)
        file_content = file.read().decode('utf-8')
        
        file_id = upload_file_to_drive(file_content, filename, DRIVE_CONFIGS_FOLDER_ID)
        if not file_id:
            flash('Could not save the configuration to Google Drive. '
                  'Nothing was stored, please try again.', 'error')
            return redirect(url_for('index'))
        
        # Parse configuration
        parsed = parse_configuration_xml(file_content)
        
        # Parse filename for auto-fill metadata
        meta = parse_scenario_filename(filename)
        
        # Add scenario to sheet with UUID
        scenario_id = str(uuid.uuid4())[:8]  # Short unique ID (8 chars)
        upload_date = datetime.now().isoformat()
        
        # Auto-fill from filename parsing
        project_name = meta.get('project_name', '')
        scenario_abbrev = meta.get('scenario_abbrev', '')
        date_run = meta.get('date_run', '')
        comment = meta.get('comment', '')
        person_name = meta.get('person_name', '')
        
        scenarios_sheet.append_row([
            scenario_id,
            parsed['scenario_name'],
            scenario_abbrev,  # personal_scenario_name / Other Name
            project_name,  # project_name
            date_run,  # date_run
            comment,  # description (filename comment goes here now)
            '',  # zaratan_link
            '',  # additional_notes -> now "Error notes", empty on upload
            uploaded_by,
            upload_date,
            file_id,
            len(parsed['input_files']),
            '',  # errors (col 13)
            '',  # submitted (col 14)
            '',  # finished (col 15)
            '',  # job_id (col 16)
            '',  # duration (col 17)
            person_name,  # person (col 18)
            ''   # based_on (col 19)
        ])
        
        link_config_inputs(scenario_id, parsed['input_files'], upload_date)
        
        # Build descriptive flash message
        auto_filled = []
        if project_name: auto_filled.append(f'Project: {project_name}')
        if scenario_abbrev: auto_filled.append(f'Abbrev: {scenario_abbrev}')
        if date_run: auto_filled.append(f'Date: {date_run}')
        if person_name: auto_filled.append(f'By: {person_name}')
        if comment: auto_filled.append(f'Description: {comment}')
        auto_msg = f' | Auto-filled: {", ".join(auto_filled)}' if auto_filled else ''
        
        flash(f'Successfully uploaded scenario "{parsed["scenario_name"]}" with {len(parsed["input_files"])} input files{auto_msg}', 'success')
        invalidate_cache()
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/upload_input', methods=['POST'])
def upload_input():
    """Upload input file XML"""
    if 'input_file' not in request.files:
        flash('No file provided', 'error')
        return redirect(url_for('index'))
    
    file = request.files['input_file']
    uploaded_by = request.form.get('uploaded_by', 'Unknown')
    
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('index'))
    
    if not allowed_file(file.filename):
        flash('Only XML files allowed', 'error')
        return redirect(url_for('index'))
    
    try:
        # Read file content
        filename = secure_filename(file.filename)
        file_content = file.read().decode('utf-8')
        
        # Parse input file (but don't store the content - just metadata)
        parsed = parse_input_file_xml(file_content)
        
        upload_date = datetime.now().isoformat()
        
        # Check if file already exists
        existing_row = find_row_by_value(inputs_sheet, 'file_name', filename)
        
        if existing_row:
            # Update existing record (no file storage)
            inputs_sheet.update_cell(existing_row, 3, parsed['regions'])
            inputs_sheet.update_cell(existing_row, 4, parsed['years'])
            inputs_sheet.update_cell(existing_row, 5, parsed['sectors'])
            inputs_sheet.update_cell(existing_row, 10, uploaded_by)
            inputs_sheet.update_cell(existing_row, 11, upload_date)
            # Leave file_id empty for input files
            flash(f'Updated input file metadata "{filename}" (file not stored)', 'success')
        else:
            # Create new record (no file storage)
            input_id = get_next_id(inputs_sheet)
            inputs_sheet.append_row([
                input_id,
                filename,
                parsed['regions'],
                parsed['years'],
                parsed['sectors'],
                '',  # policy_name
                '',  # folder_location  
                '',  # description
                '',  # additional_notes
                uploaded_by,
                upload_date,
                ''   # file_id (empty - not storing file)
            ])
            flash(f'Added input file metadata "{filename}" (file not stored)', 'success')
        
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/update_scenario/<scenario_id>', methods=['POST'])
def update_scenario(scenario_id):
    """Update scenario metadata"""
    try:
        row = find_row_by_value(scenarios_sheet, 'id', scenario_id)
        if not row:
            return jsonify({'status': 'error', 'message': 'Scenario not found'}), 404
        
        data = request.form

        SCENARIO_COLUMNS = {
            'personal_scenario_name': 3,
            'project_name': 4,
            'date_run': 5,
            'description': 6,
            'zaratan_link': 7,
            'additional_notes': 8,
            'errors': 13,
            'submitted': 14,
            'finished': 15,
            'job_id': 16,
            'person': 18,
            'based_on': 19,
        }

        applied = {}
        for field, column in SCENARIO_COLUMNS.items():
            if field in data:
                scenarios_sheet.update_cell(row, column, data[field])
                applied[field] = data[field]

        cached = patch_cached_record('scenarios', scenario_id, applied)

        if 'submitted' in applied or 'finished' in applied:
            try:
                if cached is not None:
                    submitted = str(cached.get('submitted', ''))
                    finished = str(cached.get('finished', ''))
                else:
                    row_data = scenarios_sheet.row_values(row)
                    submitted = row_data[13] if len(row_data) > 13 else ''
                    finished = row_data[14] if len(row_data) > 14 else ''
                if submitted and finished:
                    duration = compute_duration(submitted, finished)
                    scenarios_sheet.update_cell(row, 17, duration)
                    if cached is not None:
                        cached['duration'] = duration
            except Exception as e:
                app.logger.warning('duration recompute failed for %s: %s', scenario_id, e)

        if cached is None:
            invalidate_cache()

        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/update_input/<int:input_id>', methods=['POST'])
def update_input(input_id):
    """Update input file metadata"""
    try:
        row = find_row_by_id(inputs_sheet, input_id)
        if not row:
            return jsonify({'status': 'error', 'message': 'Input file not found'}), 404
        
        data = request.form

        INPUT_COLUMNS = {
            'policy_name': 6,
            'folder_location': 7,
            'description': 8,
            'additional_notes': 9,
        }

        applied = {}
        for field, column in INPUT_COLUMNS.items():
            if field in data:
                inputs_sheet.update_cell(row, column, data[field])
                applied[field] = data[field]

        if patch_cached_record('inputs', input_id, applied) is None:
            invalidate_cache()

        return jsonify({'status': 'success'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/scenario/<scenario_id>')
def scenario_detail(scenario_id):
    """View scenario details"""
    scenario = get_scenario_by_id(scenario_id)
    
    if not scenario:
        flash('Scenario not found', 'error')
        return redirect(url_for('index'))
    
    input_files = get_input_files_for_scenario(scenario_id)
    
    return render_template('scenario_detail.html', scenario=scenario, input_files=input_files)

@app.route('/input/<int:input_id>')
def input_detail(input_id):
    """View input file details"""
    input_file = get_input_by_id(input_id)
    
    if not input_file:
        flash('Input file not found', 'error')
        return redirect(url_for('index'))
    
    scenarios = get_scenarios_for_input(input_id)
    
    return render_template('input_detail.html', input_file=input_file, scenarios=scenarios)

@app.route('/compare/<int:id1>/<int:id2>')
def compare_inputs(id1, id2):
    """Compare two input files"""
    file1 = get_input_by_id(id1)
    file2 = get_input_by_id(id2)
    
    if not file1 or not file2:
        flash('One or both files not found', 'error')
        return redirect(url_for('index'))
    
    # Note: Input files are not stored, only metadata
    # Comparison not available for input files
    diff_html = None
    
    return render_template('compare.html', 
                         file1=file1, 
                         file2=file2, 
                         diff_html=diff_html,
                         message="Input file contents are not stored. Only metadata is tracked.")

@app.route('/download/<path:file_type>/<file_id>')
def download_file(file_type, file_id):
    """Download a file from Google Sheets"""
    try:
        if file_type == 'config':
            scenario = get_scenario_by_id(file_id)
            if not scenario or not scenario.get('config_file_id'):
                flash('File not found', 'error')
                return redirect(url_for('index'))
            
            sheet_file_id = scenario['config_file_id']
            filename = f"{scenario['scenario_name']}.xml"
            
           
            content = get_file_content(sheet_file_id)
            
            if not content:
                flash('Error downloading file', 'error')
                return redirect(url_for('index'))
            
            # Return file
            return Response(
                content,
                mimetype='application/xml',
                headers={'Content-Disposition': f'attachment;filename={filename}'}
            )
        
        elif file_type == 'input':
            # Input files are not stored
            flash('Input files are not stored. Only metadata is tracked.', 'info')
            return redirect(url_for('index'))
        else:
            flash('Invalid file type', 'error')
            return redirect(url_for('index'))
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/migrate_folder_locations')
def migrate_folder_locations():
    """One-time migration to populate folder_location for existing input files"""
    if not sheets_available():
        return jsonify({'error': 'Sheets not available'}), 503
    
    try:
        # Get all scenarios with their config files
        scenarios = read_scenario_records()
        input_records = read_input_records()
        
        # Build a complete map of filename -> folder_location from all configs
        file_folder_map = {}
        
        for scenario in scenarios:
            config_file_id = scenario.get('config_file_id')
            if not config_file_id:
                continue
            
            # Get the config file content
            config_content = get_file_content(config_file_id)
            if not config_content:
                continue
            
            # Parse it
            try:
                parsed = parse_configuration_xml(config_content)
                
                # Add to map (later configs will override if same filename)
                for input_file in parsed['input_files']:
                    file_folder_map[input_file['file_name']] = input_file.get('folder_location', '')
            except Exception as e:
                print(f"Error parsing config for scenario {scenario.get('id')}: {e}")
                continue
        
        # Now update all input files using individual updates (slower but safer)
        updated_count = 0
        
        for i, input_rec in enumerate(input_records, start=2):  # Start at row 2 (after header)
            file_name = input_rec['file_name']
            if file_name in file_folder_map:
                folder_loc = file_folder_map[file_name]
                # Update column 7 (folder_location)
                try:
                    inputs_sheet.update_cell(i, 7, folder_loc)
                    updated_count += 1
                    
                    # Add small delay every 50 updates to avoid rate limits
                    if updated_count % 50 == 0:
                        print(f"Updated {updated_count} files, pausing briefly...")
                        import time
                        time.sleep(2)  # 2 second pause
                except Exception as e:
                    print(f"Error updating row {i}: {e}")
                    continue
        
        return jsonify({
            'status': 'success',
            'message': f'Updated folder locations for {updated_count} input files'
        })
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Migration error: {error_details}")
        return jsonify({'error': str(e), 'details': error_details}), 500

@app.route('/test_compare')
def test_compare():
    """Test endpoint to verify routing works"""
    return jsonify({'status': 'ok', 'message': 'Comparison routing is working!'})

@app.route('/cleanup_orphaned_junctions')
def cleanup_orphaned_junctions():
    """Sweep out ALL orphaned data: junctions pointing to deleted scenarios,
    and input files no longer referenced by any surviving scenario.
    
    Uses fast read-filter-rewrite (no per-row deletes, no rate limits).
    Safe to run anytime.
    """
    if not sheets_available():
        return jsonify({'error': 'Sheets not available'}), 503
    try:
        # Valid scenario IDs (the ones that still exist)
        scenarios = read_scenario_records()
        valid_scenario_ids = set(str(s.get('id', '')).strip() for s in scenarios)
        
        # --- Clean junctions ---
        junction_values = junction_sheet.get_all_values()
        junctions_removed = 0
        still_used_input_ids = set()
        if junction_values:
            j_header = junction_values[0]
            j_data = junction_values[1:]
            try:
                sid_idx = j_header.index('scenario_id')
                iid_idx = j_header.index('input_file_id')
            except ValueError:
                sid_idx, iid_idx = 0, 1
            
            surviving_junctions = []
            for row in j_data:
                if len(row) <= max(sid_idx, iid_idx):
                    continue
                if str(row[sid_idx]).strip() in valid_scenario_ids:
                    surviving_junctions.append(row)
                    still_used_input_ids.add(str(row[iid_idx]).strip())
                else:
                    junctions_removed += 1
            
            if junctions_removed > 0:
                junction_sheet.clear()
                junction_sheet.append_row(j_header)
                if surviving_junctions:
                    junction_sheet.append_rows(surviving_junctions)
        
        # --- Clean orphaned input files (not referenced by any surviving junction) ---
        input_values = inputs_sheet.get_all_values()
        inputs_removed = 0
        if input_values:
            i_header = input_values[0]
            i_data = input_values[1:]
            try:
                id_idx = i_header.index('id')
            except ValueError:
                id_idx = 0
            
            surviving_inputs = []
            for row in i_data:
                if len(row) <= id_idx:
                    continue
                if str(row[id_idx]).strip() in still_used_input_ids:
                    surviving_inputs.append(row)
                else:
                    inputs_removed += 1
            
            if inputs_removed > 0:
                inputs_sheet.clear()
                inputs_sheet.append_row(i_header)
                if surviving_inputs:
                    inputs_sheet.append_rows(surviving_inputs)
        
        invalidate_cache()
        return jsonify({
            'status': 'success',
            'orphaned_junctions_removed': junctions_removed,
            'orphaned_input_files_removed': inputs_removed,
            'valid_scenarios': len(valid_scenario_ids)
        })
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@app.route('/add_zaratan_columns')
def add_zaratan_columns():
    """One-time migration to add new Zaratan columns to Scenarios sheet"""
    if not sheets_available():
        return jsonify({'error': 'Sheets not available'}), 503
    try:
        headers = scenarios_sheet.row_values(1)
        new_cols = ['errors', 'submitted', 'finished', 'job_id', 'duration', 'person', 'based_on', 'zaratan_username']
        added = []
        for col in new_cols:
            if col not in headers:
                next_col = len(headers) + 1
                scenarios_sheet.update_cell(1, next_col, col)
                headers.append(col)
                added.append(col)
        return jsonify({'status': 'success', 'added_columns': added, 'all_headers': headers})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/compare_scenarios', strict_slashes=False)
def compare_scenarios():
    """Compare multiple scenarios and generate a report"""
    try:
        print("DEBUG: Compare scenarios route called")
        print(f"DEBUG: Request args: {request.args}")
        scenario_ids = request.args.get('ids', '').split(',')
        print(f"DEBUG: Scenario IDs: {scenario_ids}")
        
        if len(scenario_ids) < 2:
            flash('Please select at least 2 scenarios to compare', 'error')
            return redirect(url_for('index'))
        
        if not sheets_available():
            flash('Google Sheets connection unavailable', 'error')
            return redirect(url_for('index'))
        
        # Get all scenarios
        scenarios = []
        for scenario_id in scenario_ids:
            print(f"DEBUG: Getting scenario {scenario_id}")
            scenario = get_scenario_by_id(scenario_id)
            if scenario:
                # Get input files for this scenario
                input_files = get_input_files_for_scenario(scenario_id)
                scenario['input_file_names'] = set([f['file_name'] for f in input_files])
                scenarios.append(scenario)
                print(f"DEBUG: Found scenario with {len(input_files)} input files")
            else:
                print(f"DEBUG: Scenario {scenario_id} not found")
        
        if len(scenarios) < 2:
            flash('Could not find all selected scenarios', 'error')
            return redirect(url_for('index'))
        
        # Generate comparison report
        report_lines = []
        report_lines.append("=" * 80)
        report_lines.append("GCAM SCENARIO COMPARISON REPORT")
        report_lines.append("=" * 80)
        report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report_lines.append(f"Number of scenarios compared: {len(scenarios)}")
        report_lines.append("")
        
        # Section 1: Scenario Overview
        report_lines.append("-" * 80)
        report_lines.append("SCENARIO OVERVIEW")
        report_lines.append("-" * 80)
        for i, scenario in enumerate(scenarios, 1):
            report_lines.append(f"\n{i}. {scenario['scenario_name']}")
            report_lines.append(f"   ID: {scenario['id']}")
            if scenario.get('project_name'):
                report_lines.append(f"   Project: {scenario['project_name']}")
            if scenario.get('date_run'):
                report_lines.append(f"   Date Run: {scenario['date_run']}")
            report_lines.append(f"   Number of Input Files: {len(scenario['input_file_names'])}")
            if scenario.get('description'):
                report_lines.append(f"   Description: {scenario['description']}")
        
        report_lines.append("")
        
        # Section 2: Input Files Comparison - UNIQUE FILES FIRST
        report_lines.append("-" * 80)
        report_lines.append("INPUT FILES COMPARISON")
        report_lines.append("-" * 80)
        
        # Get all unique files across all scenarios
        all_files = set()
        for scenario in scenarios:
            all_files.update(scenario['input_file_names'])
        
        # Files shared by ALL scenarios
        shared_files = set.intersection(*[s['input_file_names'] for s in scenarios])
        
        report_lines.append(f"\nTotal unique input files across all scenarios: {len(all_files)}")
        report_lines.append(f"Files shared by ALL scenarios: {len(shared_files)}")
        
        # UNIQUE FILES FIRST (moved up)
        report_lines.append("\n" + "-" * 80)
        report_lines.append("UNIQUE FILES PER SCENARIO")
        report_lines.append("-" * 80)
        
        for scenario in scenarios:
            unique_files = scenario['input_file_names'] - shared_files
            other_files = set()
            for other in scenarios:
                if other['id'] != scenario['id']:
                    other_files.update(other['input_file_names'])
            
            truly_unique = scenario['input_file_names'] - other_files
            
            report_lines.append(f"\n{scenario['scenario_name']}:")
            report_lines.append(f"  Files NOT in common set: {len(unique_files)}")
            report_lines.append(f"  Files ONLY in this scenario: {len(truly_unique)}")
            
            if truly_unique:
                report_lines.append("  Unique files:")
                for file in sorted(truly_unique):
                    report_lines.append(f"    - {file}")
        
        # SHARED FILES AFTER (moved down)
        report_lines.append("\n" + "-" * 80)
        report_lines.append("SHARED FILES")
        report_lines.append("-" * 80)
        
        if shared_files:
            report_lines.append(f"\nFiles present in ALL {len(scenarios)} scenarios ({len(shared_files)} total):")
            for file in sorted(shared_files):
                report_lines.append(f"  - {file}")
        else:
            report_lines.append("\nNo files are shared by all scenarios.")
        
        # Section 3: File-by-File Matrix
        report_lines.append("\n" + "-" * 80)
        report_lines.append("FILE PRESENCE MATRIX")
        report_lines.append("-" * 80)
        report_lines.append("\nLegend: ✓ = Present, ✗ = Absent\n")
        
        # Create header
        header = "File Name".ljust(50)
        for i, scenario in enumerate(scenarios, 1):
            header += f"  S{i}"
        report_lines.append(header)
        report_lines.append("-" * len(header))
        
        # Add each file
        for file in sorted(all_files):
            line = file[:48].ljust(50)
            for scenario in scenarios:
                if file in scenario['input_file_names']:
                    line += "  ✓ "
                else:
                    line += "  ✗ "
            report_lines.append(line)
        
        # Section 4: Summary Statistics
        report_lines.append("\n" + "-" * 80)
        report_lines.append("SUMMARY STATISTICS")
        report_lines.append("-" * 80)
        
        for i, scenario in enumerate(scenarios, 1):
            overlap_counts = []
            for j, other in enumerate(scenarios, 1):
                if i != j:
                    overlap = len(scenario['input_file_names'] & other['input_file_names'])
                    total = len(scenario['input_file_names'] | other['input_file_names'])
                    if total > 0:
                        percentage = (overlap / total) * 100
                        overlap_counts.append(f"S{j}: {percentage:.1f}%")
            
            report_lines.append(f"\nS{i} ({scenario['scenario_name']}) overlap with others:")
            report_lines.append(f"  {', '.join(overlap_counts)}")
        
        report_lines.append("\n" + "=" * 80)
        report_lines.append("END OF REPORT")
        report_lines.append("=" * 80)
        
        # =====================================================================
        # Generate XLSX comparison spreadsheet (2 sheets)
        # =====================================================================
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        
        # Build file -> folder map and file -> scenarios map
        file_folder_map = {}
        file_scenario_map = {}
        
        # Get folder info from input files sheet
        try:
            all_inputs = read_input_records()
            for inp in all_inputs:
                file_folder_map[inp.get('file_name', '')] = inp.get('folder_location', '')
        except:
            pass
        
        scenario_names_list = [s['scenario_name'] for s in scenarios]
        
        for scenario in scenarios:
            for file_name in scenario['input_file_names']:
                if file_name not in file_scenario_map:
                    file_scenario_map[file_name] = set()
                file_scenario_map[file_name].add(scenario['scenario_name'])
        
        all_files_sorted = sorted(file_scenario_map.keys())
        total_scenarios = len(scenarios)
        
        # Styles - only for headers
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="991B1B")
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # --- Sheet 1: Summary (first tab) ---
        ws1 = wb.active
        ws1.title = "Summary"
        
        summary_headers = ["File Name", "Folder", "Status", "Present In"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        current_row = 2
        
        # Unique files first
        for file_name in all_files_sorted:
            present_in = file_scenario_map.get(file_name, set())
            if len(present_in) < total_scenarios:
                folder = file_folder_map.get(file_name, '')
                
                if len(present_in) == 1:
                    status = f"Unique to {list(present_in)[0]}"
                    present_str = ''  # Leave empty for unique files
                else:
                    status = f"In {len(present_in)} of {total_scenarios} scenarios"
                    present_str = "; ".join(sorted(present_in))
                
                ws1.cell(row=current_row, column=1, value=file_name).border = border
                ws1.cell(row=current_row, column=2, value=folder).border = border
                ws1.cell(row=current_row, column=3, value=status).border = border
                ws1.cell(row=current_row, column=4, value=present_str).border = border
                current_row += 1
        
        # Separator row
        sep_cell = ws1.cell(row=current_row, column=1, value=f"--- SHARED FILES ({len(shared_files)} files in all {total_scenarios} scenarios) ---")
        sep_cell.font = Font(bold=True)
        current_row += 1
        
        # Shared files
        for file_name in sorted(shared_files):
            folder = file_folder_map.get(file_name, '')
            present_str = "; ".join(scenario_names_list)
            
            ws1.cell(row=current_row, column=1, value=file_name).border = border
            ws1.cell(row=current_row, column=2, value=folder).border = border
            ws1.cell(row=current_row, column=3, value="Shared by all").border = border
            ws1.cell(row=current_row, column=4, value=present_str).border = border
            current_row += 1
        
        ws1.column_dimensions['A'].width = 40
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 50
        ws1.freeze_panes = 'A2'
        
        # --- Sheet 2: Presence Matrix ---
        ws2 = wb.create_sheet("Presence Matrix")
        
        headers = ["File Name", "Folder"] + scenario_names_list
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Data rows - no fill colors, just 1/0
        for row_idx, file_name in enumerate(all_files_sorted, 2):
            folder = file_folder_map.get(file_name, '')
            present_in = file_scenario_map.get(file_name, set())
            
            ws2.cell(row=row_idx, column=1, value=file_name).border = border
            ws2.cell(row=row_idx, column=2, value=folder).border = border
            
            for sc_idx, sc_name in enumerate(scenario_names_list, 3):
                is_present = sc_name in present_in
                cell = ws2.cell(row=row_idx, column=sc_idx, value=1 if is_present else 0)
                cell.alignment = center_align
                cell.border = border
        
        # Summary row
        summary_row = len(all_files_sorted) + 3
        ws2.cell(row=summary_row, column=1, value="TOTAL FILES").font = Font(bold=True)
        for sc_idx, sc_name in enumerate(scenario_names_list, 3):
            count = sum(1 for f in all_files_sorted if sc_name in file_scenario_map.get(f, set()))
            cell = ws2.cell(row=summary_row, column=sc_idx, value=count)
            cell.font = Font(bold=True)
            cell.alignment = center_align
        
        ws2.column_dimensions['A'].width = 40
        ws2.column_dimensions['B'].width = 20
        from openpyxl.utils import get_column_letter
        for i in range(3, 3 + len(scenario_names_list)):
            ws2.column_dimensions[get_column_letter(i)].width = max(15, len(scenario_names_list[i-3]) + 4)
        ws2.freeze_panes = 'A2'
        
        # Save
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        return Response(
            xlsx_buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment;filename=scenario_comparison_{timestamp}.xlsx'}
        )
        
    except Exception as e:
        import traceback
        print(f"Comparison error: {traceback.format_exc()}")
        flash(f'Error comparing scenarios: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/delete_scenario/<scenario_id>', methods=['POST'])
def delete_scenario(scenario_id):
    """Delete a scenario, its junctions, and any input files left orphaned.
    
    Uses a read-filter-rewrite approach (clear + rewrite) instead of deleting
    rows one by one, so it stays fast and avoids Google's write rate limits.
    Input files still used by OTHER scenarios are preserved.
    """
    try:
        if not sheets_available():
            return jsonify({'success': False, 'error': 'Sheets unavailable'})
        
        scenario_row = find_row_by_value(scenarios_sheet, 'id', scenario_id)
        if not scenario_row:
            return jsonify({'success': False, 'error': 'Scenario not found'})
        
        sid = str(scenario_id).strip()
        
        # 1. Delete config file from FileStorage
        scenario_data = scenarios_sheet.row_values(scenario_row)
        config_file_id = scenario_data[10] if len(scenario_data) > 10 else None
        if config_file_id:
            try:
                file_row = find_row_by_value(file_storage_sheet, 'file_id', config_file_id)
                if file_row:
                    file_storage_sheet.delete_rows(file_row)
            except Exception as e:
                print(f"Error deleting config from storage: {e}")
        
        # 2. Delete the scenario row
        scenarios_sheet.delete_rows(scenario_row)
        print(f"Deleted scenario {sid}")
        
        # 3. Process junctions: figure out which input files this scenario used,
        #    which junctions survive, and which input files become orphaned.
        junction_values = junction_sheet.get_all_values()
        orphaned_input_ids = set()
        if junction_values:
            j_header = junction_values[0]
            j_data = junction_values[1:]
            try:
                sid_idx = j_header.index('scenario_id')
                iid_idx = j_header.index('input_file_id')
            except ValueError:
                sid_idx, iid_idx = 0, 1  # fallback to known order
            
            this_scenario_inputs = set()
            surviving_junctions = []
            for row in j_data:
                if len(row) <= max(sid_idx, iid_idx):
                    continue
                row_sid = str(row[sid_idx]).strip()
                row_iid = str(row[iid_idx]).strip()
                if row_sid == sid:
                    this_scenario_inputs.add(row_iid)
                else:
                    surviving_junctions.append(row)
            
            # Which of this scenario's input files are still used by others?
            still_used = set(str(r[iid_idx]).strip() for r in surviving_junctions if len(r) > iid_idx)
            orphaned_input_ids = this_scenario_inputs - still_used
            
            # Rewrite junction sheet without this scenario's junctions
            junction_sheet.clear()
            junction_sheet.append_row(j_header)
            if surviving_junctions:
                junction_sheet.append_rows(surviving_junctions)
            print(f"Junctions: removed {len(this_scenario_inputs)} links, {len(surviving_junctions)} remain")
        
        # 4. Delete orphaned input files (used only by this scenario)
        if orphaned_input_ids:
            input_values = inputs_sheet.get_all_values()
            if input_values:
                i_header = input_values[0]
                i_data = input_values[1:]
                try:
                    id_idx = i_header.index('id')
                except ValueError:
                    id_idx = 0
                
                surviving_inputs = [
                    row for row in i_data
                    if len(row) > id_idx and str(row[id_idx]).strip() not in orphaned_input_ids
                ]
                
                inputs_sheet.clear()
                inputs_sheet.append_row(i_header)
                if surviving_inputs:
                    inputs_sheet.append_rows(surviving_inputs)
                print(f"Input files: deleted {len(orphaned_input_ids)} orphaned, {len(surviving_inputs)} remain")
        
        invalidate_cache()
        return jsonify({'success': True})
        
    except Exception as e:
        import traceback
        print(f"Error deleting scenario: {traceback.format_exc()}")
        invalidate_cache()
        return jsonify({'success': False, 'error': str(e)})

# =============================================================================
# Zaratan Log Ingestion
# =============================================================================
def zaratan_log_to_row(record):
    """Flatten a single Zaratan log record (started or finished) into a row
    matching ZARATAN_LOG_COLUMNS. Missing fields become empty strings and the
    full original record is preserved in raw_json."""
    sacct = record.get('sacct') or {}
    values = {
        'job_id': record.get('job_id', ''),
        'restart_count': record.get('restart_count', ''),
        'event': record.get('event', ''),
        'scenario_name': record.get('scenario_name', ''),
        'project_name': record.get('project_name', ''),
        'user': record.get('user', ''),
        'hostname': record.get('hostname', ''),
        'started_at': record.get('started_at', ''),
        'config_file': record.get('config_file', ''),
        'gcam_dir': record.get('gcam_dir', ''),
        'gcamreport_template_version': record.get('gcamreport_template_version', ''),
        'gcam_finished_at': record.get('gcam_finished_at', ''),
        'gcam_exit_code': record.get('gcam_exit_code', ''),
        'script_finished_at': record.get('script_finished_at', ''),
        'script_exit_code': record.get('script_exit_code', ''),
        'sacct_submit': sacct.get('submit', ''),
        'sacct_start': sacct.get('start', ''),
        'sacct_end': sacct.get('end', ''),
        'sacct_elapsed': sacct.get('elapsed', ''),
        'sacct_state': sacct.get('state', ''),
        'sacct_exit_code': sacct.get('exit_code', ''),
        'received_at': datetime.now().isoformat(),
        'raw_json': json.dumps(record, ensure_ascii=False),
    }
    return [values[col] for col in ZARATAN_LOG_COLUMNS]

def zaratan_log_key(job_id, restart_count, event):
    """Natural unique key for a log record: one started/finished event per job
    attempt. Normalized to strings so values from the sheet (always strings)
    compare equal to values from JSON (job_id str, restart_count int)."""
    return (str(job_id).strip(), str(restart_count).strip(), str(event).strip())

def _log_date_run(record):
    """YYYY-MM-DD the run actually ran, from started_at (or sacct.start)."""
    ts = record.get('started_at') or (record.get('sacct') or {}).get('start') or ''
    ts = str(ts)
    return ts[:10] if len(ts) >= 10 and ts[:4].isdigit() else ''

def _log_finished(record):
    """Authoritative finish time: sacct.end unless it's still Unknown, in which
    case fall back to the script-written gcam_finished_at."""
    end = str((record.get('sacct') or {}).get('end', '')).strip()
    if end and end.lower() != 'unknown':
        return end
    return str(record.get('gcam_finished_at', '') or '')

def _log_errors(record):
    """Whether this run errored, as a Yes/No flag to match the webapp's current
    'errors' checkbox (truthy = checked). A run errors if SLURM ended in a
    non-success state or GCAM returned a non-zero exit code.

    The detailed reason is preserved (commented out below) for when we want to
    surface it, e.g. routed into the "Error notes" column. The full detail always
    lives in ZaratanLogs regardless.
    """
    sacct = record.get('sacct') or {}
    state = str(sacct.get('state', '')).strip()
    gcam_ec = str(record.get('gcam_exit_code', '')).strip()

    # --- Detailed reason (kept for later use) ---
    # problems = []
    # if state and state.upper() not in ('COMPLETED', 'RUNNING'):
    #     problems.append(state)
    # if gcam_ec and gcam_ec not in ('0', ''):
    #     problems.append(f'gcam_exit={gcam_ec}')
    # return '; '.join(problems)

    bad_state = state.upper() not in ('', 'COMPLETED', 'RUNNING', 'PENDING')
    bad_exit = gcam_ec not in ('', '0')
    return 'Yes' if (bad_state or bad_exit) else ''

def _log_person(meta, user):
    """Resolve the runner's full name for a logged run. Tries the -FL initials
    parsed from scenario_name (meta['person_name'] via PERSON_MAP) first, then
    falls back to the Zaratan username mapping (USERNAME_MAP). Returns '' if the
    person can't be identified from either source. `meta` is the result of
    parse_scenario_filename(scenario_name)."""
    name = meta.get('person_name', '')
    if name:
        return name
    if user:
        return USERNAME_MAP.get(user, '')
    return ''

def scenario_fields_from_log(record):
    """The Scenarios columns a single log record should set (empty values omitted
    so they never overwrite existing data). started supplies identity + date_run;
    finished supplies submitted/finished/duration/errors."""
    f = {}
    scenario_name = record.get('scenario_name', '')
    meta = parse_scenario_filename(scenario_name) if scenario_name else {}
    if scenario_name:
        f['scenario_name'] = scenario_name
    code = record.get('project_name')
    if code:
        f['project_name'] = PROJECT_MAP.get(code, code)
    if record.get('job_id'):
        f['job_id'] = str(record['job_id'])
    user = record.get('user', '')
    if user:
        f['zaratan_username'] = user 
    person = _log_person(meta, user)
    if person:
        f['person'] = person  
    if meta.get('comment'):
        f['description'] = meta['comment']
    date_run = _log_date_run(record)
    if date_run:
        f['date_run'] = date_run
    sacct = record.get('sacct') or {}
    if sacct or record.get('event') == 'finished':
        if sacct.get('submit'):
            f['submitted'] = sacct['submit']
        finished = _log_finished(record)
        if finished:
            f['finished'] = finished
        if sacct.get('elapsed'):
            f['duration'] = sacct['elapsed']
        errors = _log_errors(record)
        if errors:
            f['errors'] = errors
    return f

def apply_config_from_logs(records, scenario_ids):
    """For log records that carry the run's configuration XML, archive it to
    Drive and link the input files it declares to the scenario.

    This is what gives Zaratan runs the same input-file listing that a website
    upload produces. Without it the tracker knows a run happened but not what
    went into it.

    Skips a scenario that already has a config recorded, so re-sending a log
    never duplicates junction rows. Returns a summary dict.
    """
    if scenarios_sheet is None:
        return {'configs': 0, 'linked': 0, 'skipped': 0, 'errors': 0}

    header = scenarios_sheet.row_values(1)
    try:
        col_config = header.index('config_file_id') + 1
        id_idx = header.index('id')
        cfg_idx = header.index('config_file_id')
    except ValueError:
        return {'configs': 0, 'linked': 0, 'skipped': 0, 'errors': 0}

    # Which scenarios already have a config, and where their row is
    rows = scenarios_sheet.get_all_values()
    row_of, has_config = {}, set()
    for n, row in enumerate(rows[1:], start=2):
        if len(row) > id_idx:
            sid = str(row[id_idx]).strip()
            if sid:
                row_of[sid] = n
                if len(row) > cfg_idx and str(row[cfg_idx]).strip():
                    has_config.add(sid)

    summary = {'configs': 0, 'linked': 0, 'skipped': 0, 'errors': 0}
    updates = []
    upload_date = datetime.now().isoformat()

    # One config per job_id; a restart re-sends the same one
    seen_jobs = set()
    for rec in records:
        content = rec.get('config_content')
        jid = str(rec.get('job_id', '')).strip()
        if not content or not jid or jid in seen_jobs:
            continue
        seen_jobs.add(jid)

        scenario_id = scenario_ids.get(jid)
        if not scenario_id:
            summary['skipped'] += 1
            continue
        if scenario_id in has_config:
            summary['skipped'] += 1
            continue

        try:
            parsed = parse_configuration_xml(content)
        except Exception as e:
            print(f"WARN: could not parse config for job {jid}: {e}")
            summary['errors'] += 1
            continue

        filename = rec.get('config_file') or f'configuration_{jid}.xml'
        filename = os.path.basename(filename)
        file_id = upload_file_to_drive(content, filename, DRIVE_CONFIGS_FOLDER_ID)
        if not file_id:
            print(f"WARN: Drive upload failed for job {jid}; not linking inputs")
            summary['errors'] += 1
            continue

        result = link_config_inputs(scenario_id, parsed.get('input_files', []), upload_date)
        summary['configs'] += 1
        summary['linked'] += result['linked']

        row_num = row_of.get(scenario_id)
        if row_num:
            updates.append({'range': gspread.utils.rowcol_to_a1(row_num, col_config),
                            'values': [[file_id]]})

    if updates:
        scenarios_sheet.batch_update(updates)

    return summary


# Written when the cluster parser ran but could not determine the outcome.
SOLVE_UNKNOWN_NOTE = 'Solve status unknown'


def _fmt_list(values, limit=8):
    """Comma-joined list, abbreviated once it gets long enough to be unreadable."""
    vals = [str(v) for v in values]
    if len(vals) <= limit:
        return ', '.join(vals)
    return ', '.join(vals[:limit]) + f' (+{len(vals) - limit} more)'


def solve_note(record):
    """Error-notes text derived from a log record's `solve` block.

    Returns '' when the record carries no solve information at all, so nothing
    is written for logs predating the cluster-side parser.

    'unknown' is rendered explicitly rather than left blank: a run whose log
    could not be read must never be indistinguishable from one that solved.
    """
    solve = record.get('solve')
    if not isinstance(solve, dict):
        return ''

    status = str(solve.get('status', '')).strip().lower()
    total = solve.get('periods_total')

    if status == 'solved':
        return f'All {total} periods solved' if total else 'Solved'

    if status == 'failed':
        years = solve.get('years_failed') or []
        periods = solve.get('periods_failed') or []

        if years:
            which = _fmt_list(years)
        elif periods:
            which = 'period ' + _fmt_list(periods)
        else:
            which = ''

        count = len(years) or len(periods)
        if count and total:
            head = f'{count} of {total} periods failed'
        elif count:
            head = f'{count} period{"s" if count != 1 else ""} failed'
        else:
            head = "Didn't solve"

        note = f'{head}: {which}' if which else head

        # Calibration is only worth calling out when it is a different story
        # from the solve failures - usually the two sets coincide.
        uncal = solve.get('periods_uncalibrated') or []
        if uncal and sorted(uncal) != sorted(periods):
            note += '; calibration failed in period ' + _fmt_list(uncal)

        return note

    if status == 'unknown':
        return SOLVE_UNKNOWN_NOTE

    # Unrecognised status - say nothing rather than guess
    return ''


def apply_scenario_upserts(records):
    """Reflect ingested log records into the Scenarios sheet, keyed by job_id.
    'started' creates/populates the row; 'finished' updates the same row in place.
    Restarts and finished-before-started are handled naturally by upserting on
    job_id. Returns {'created': N, 'updated': M}. Reads the sheet once and writes
    in batched calls to stay within API rate limits."""
    if scenarios_sheet is None:
        return {'created': 0, 'updated': 0}
    from gspread.utils import rowcol_to_a1

    values = scenarios_sheet.get_all_values()
    if not values:
        return {'created': 0, 'updated': 0}
    header = values[0]
    col = {name: i for i, name in enumerate(header)}
    if 'job_id' not in col:
        return {'created': 0, 'updated': 0}
    jid_idx = col['job_id']

    # Existing job_id -> sheet row number (1-indexed; data starts at row 2).
    # Rows with an empty job_id (e.g. website config uploads) are ignored, so a
    # log always gets its own row rather than attaching to an unrelated one.
    existing_rownum = {}
    existing_scenario_id = {}   # job_id -> scenario id, for linking config inputs
    id_idx = col.get('id')
    for i, row in enumerate(values[1:], start=2):
        if len(row) > jid_idx:
            jid = str(row[jid_idx]).strip()
            if jid:
                existing_rownum[jid] = i
                if id_idx is not None and len(row) > id_idx:
                    existing_scenario_id[jid] = str(row[id_idx]).strip()

    pending_new = {}      # job_id -> full row list (appended at the end)
    updates = []          # (row_number, col_idx0, value) for existing rows

    notes_idx = col.get('additional_notes')

    for rec in records:
        jid = str(rec.get('job_id', '')).strip()
        if not jid:
            continue
        fields = scenario_fields_from_log(rec)

        # Error notes are deliberately kept out of `fields`: everything in
        # there overwrites unconditionally, and a note someone typed by hand
        # must survive. 
        note = solve_note(rec)

        if jid in existing_rownum:
            rownum = existing_rownum[jid]
            for name, val in fields.items():
                if val != '' and name in col:
                    updates.append((rownum, col[name], val))

            if note and notes_idx is not None:
                current = ''
                sheet_row = values[rownum - 1] if rownum - 1 < len(values) else []
                if len(sheet_row) > notes_idx:
                    current = str(sheet_row[notes_idx]).strip()
                # Fill only when empty, or when replacing our own placeholder
                if current == '' or current == SOLVE_UNKNOWN_NOTE:
                    updates.append((rownum, notes_idx, note))
        elif jid in pending_new:
            row = pending_new[jid]
            for name, val in fields.items():
                if val != '' and name in col:
                    row[col[name]] = val
            if note and notes_idx is not None:
                existing_note = str(row[notes_idx]).strip()
                if existing_note == '' or existing_note == SOLVE_UNKNOWN_NOTE:
                    row[notes_idx] = note
        else:
            row = [''] * len(header)
            new_id = str(uuid.uuid4())[:8]
            base = {
                'id': new_id,
                # 'Zaratan' is the value the dashboard's "Uploaded through" column
                # recognizes as the cluster channel (else it shows "Website").
                'uploaded_by': 'Zaratan',
                'upload_date': datetime.now().isoformat(),
            }
            if note:
                base['additional_notes'] = note
            for name, val in {**base, **fields}.items():
                if val != '' and name in col:
                    row[col[name]] = val
            pending_new[jid] = row
            existing_scenario_id[jid] = new_id

    if pending_new:
        scenarios_sheet.append_rows(list(pending_new.values()))
    if updates:
        batch = [{'range': rowcol_to_a1(r, c + 1), 'values': [[v]]} for (r, c, v) in updates]
        scenarios_sheet.batch_update(batch)

    return {
        'created': len(pending_new),
        'updated': len(set(u[0] for u in updates)),
        # job_id -> scenario id, so the caller can attach config inputs to the
        # right scenario without re-reading the sheet
        'scenario_ids': existing_scenario_id,
    }

@app.route('/ingest_logs', methods=['POST'])
def ingest_logs():
    """Receive Zaratan run logs and append them to the ZaratanLogs worksheet.

    Accepts either a single JSON object or a JSON list of objects (a batch).
    Protected by INGEST_TOKEN when that env var is set (via ?token= or the
    X-Ingest-Token header).

    Records already present (same job_id+restart_count+event) are skipped, so
    re-sending a batch never creates duplicate rows. Newly-stored records are
    also upserted into the Scenarios sheet by job_id (started creates the row,
    finished updates it). Returns {"received": N, "skipped": M, "scenarios":
    {...}} on success; the sender can safely delete all files in the batch after
    any 200 response.
    """
    # Auth (only enforced if a token is configured)
    if INGEST_TOKEN:
        provided = request.headers.get('X-Ingest-Token') or request.args.get('token', '')
        if provided != INGEST_TOKEN:
            return jsonify({'status': 'error', 'message': 'Unauthorized'}), 401

    if zaratan_logs_sheet is None:
        return jsonify({'status': 'error', 'message': 'Sheets unavailable'}), 503

    payload = request.get_json(silent=True)
    if payload is None:
        return jsonify({'status': 'error', 'message': 'Invalid or missing JSON body'}), 400

    records = payload if isinstance(payload, list) else [payload]
    if not records:
        return jsonify({'status': 'ok', 'received': 0, 'skipped': 0})

    try:
        # Build the set of keys already stored so re-sent batches don't create
        # duplicate rows. job_id/restart_count/event are the first 3 columns.
        existing = zaratan_logs_sheet.get_all_values()
        seen = set()
        for row in existing[1:]:  # skip header
            if len(row) >= 3:
                seen.add(zaratan_log_key(row[0], row[1], row[2]))

        rows = []
        new_records = []
        skipped = 0
        for r in records:
            if not isinstance(r, dict):
                continue
            key = zaratan_log_key(r.get('job_id', ''), r.get('restart_count', ''), r.get('event', ''))
            if key in seen:  # already in the sheet, or a duplicate within this batch
                skipped += 1
                continue
            seen.add(key)
            rows.append(zaratan_log_to_row(r))
            new_records.append(r)

        if rows:
            zaratan_logs_sheet.append_rows(rows)
        scenarios = {'created': 0, 'updated': 0}
        configs = {'configs': 0, 'linked': 0, 'skipped': 0, 'errors': 0}
        scenarios_error = None
        try:
            scenarios = apply_scenario_upserts(new_records)
            # Logs that carry the run's config give the scenario its input files
            configs = apply_config_from_logs(new_records, scenarios.get('scenario_ids', {}))
            invalidate_cache()
        except Exception as se:
            scenarios_error = str(se)

        scenarios_summary = {k: v for k, v in scenarios.items() if k != 'scenario_ids'}
        result = {'status': 'ok', 'received': len(rows), 'skipped': skipped,
                  'scenarios': scenarios_summary, 'configs': configs}
        if scenarios_error:
            result['scenarios_error'] = scenarios_error
        return jsonify(result)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =============================================================================
# Main
# =============================================================================

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
